import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column names matching user request
HEADERS = [
    "S.No.",
    "School Name",
    "Customer Name",
    "Institution Type",
    "Social Media",
    "Area Name",
    "Institution Atmosphere",
    "Website URL",
    "Contact Number",
    "School Address",
    "Pincode",
    "Appearance",
    "Remarks"
]

# Color styling codes matching screenshot aesthetics
COLOR_BLUE_HEADER = "002060"    # Dark Blue header
COLOR_WHITE = "FFFFFF"          # White text
COLOR_GREEN_ACTIVE = "1F4E3D"   # Dark green for Active / Good / Redesign
COLOR_RED_INACTIVE = "9C0006"   # Dark red for Inactive / Bad
COLOR_PURPLE_FRESH = "7030A0"   # Purple for Fresh
COLOR_LIGHT_GRAY = "F2F2F2"     # Row zebra striping

def save_leads_to_excel(leads: list, filename: str = "school_leads.xlsx") -> str:
    """
    Saves the lead data into a styled Excel file that matches the client's screenshot.
    """
    try:
        # Create a new workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "School Leads"
        
        # Show gridlines
        ws.views.sheetView[0].showGridLines = True
        
        # Apply header styling
        header_font = Font(name="Arial", size=11, bold=True, color=COLOR_WHITE)
        header_fill = PatternFill(start_color=COLOR_BLUE_HEADER, end_color=COLOR_BLUE_HEADER, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.append(HEADERS)
        ws.row_dimensions[1].height = 28
        
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Borders definition
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Add lead rows
        for idx, lead in enumerate(leads, 1):
            row_data = [
                idx,  # S.No.
                lead.get("school_name", ""),
                "",   # Customer Name (Not necessary)
                lead.get("institution_type", "Matriculation"),
                lead.get("social_media", "Inactive"),
                lead.get("area_name", ""),
                lead.get("atmosphere", "Good"),
                lead.get("website_url", ""),
                lead.get("contact_number", ""),
                lead.get("address", ""),
                lead.get("pincode", ""),
                lead.get("appearance", "Redesign"),
                lead.get("remarks", "")
            ]
            ws.append(row_data)
            row_num = idx + 1
            ws.row_dimensions[row_num].height = 22
            
            # Alignments & Styles
            for col_idx in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                
                val_str = str(cell.value or "").strip()
                
                # Check column-specific styling matching the screenshot
                # Column 1: S.No (Center)
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Column 4: Institution Type (Center)
                elif col_idx == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # light blue
                    cell.font = Font(name="Arial", size=10, color="305496")
                    
                # Column 5: Social Media (Center, color indicators)
                elif col_idx == 5:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val_str == "Active":
                        cell.fill = PatternFill(start_color=COLOR_GREEN_ACTIVE, end_color=COLOR_GREEN_ACTIVE, fill_type="solid")
                        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
                    else:
                        cell.fill = PatternFill(start_color=COLOR_RED_INACTIVE, end_color=COLOR_RED_INACTIVE, fill_type="solid")
                        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
                        
                # Column 7: Institution Atmosphere (Center, color indicators)
                elif col_idx == 7:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val_str == "Good":
                        cell.fill = PatternFill(start_color=COLOR_GREEN_ACTIVE, end_color=COLOR_GREEN_ACTIVE, fill_type="solid")
                        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
                    else:
                        cell.fill = PatternFill(start_color=COLOR_RED_INACTIVE, end_color=COLOR_RED_INACTIVE, fill_type="solid")
                        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
                        
                # Column 8: Website URL (Left, Blue underlined link format)
                elif col_idx == 8 and val_str:
                    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                    
                # Column 11: Appearance (Center, color indicators)
                elif col_idx == 11:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val_str == "Redesign":
                        cell.fill = PatternFill(start_color=COLOR_GREEN_ACTIVE, end_color=COLOR_GREEN_ACTIVE, fill_type="solid")
                        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
                    elif val_str == "Fresh":
                        cell.fill = PatternFill(start_color=COLOR_PURPLE_FRESH, end_color=COLOR_PURPLE_FRESH, fill_type="solid")
                        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
                    else:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
                        cell.font = Font(name="Arial", size=10, color="375623")
                        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Avoid measuring long URLs for column widths (keep them compact)
            for cell in col:
                val = str(cell.value or "")
                if cell.column == 8:  # Website URL col
                    max_len = max(max_len, min(len(val), 25))
                else:
                    max_len = max(max_len, len(val))
                    
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(filename)
        logger.info(f"Excel report saved successfully to {filename}")
        return os.path.abspath(filename)
        
    except Exception as e:
        logger.error(f"Failed to save leads to Excel: {e}")
        return ""

def sync_to_google_sheets(leads: list, sheet_id: str) -> bool:
    """
    Placeholder/Interface implementation for Google Sheets Sync.
    Requires google-auth credentials setup. Will log instructions.
    """
    if not sheet_id:
        logger.info("No GOOGLE_SHEET_ID set. Skipping Google Sheets sync.")
        return False
        
    credentials_file = "credentials.json"
    if not os.path.exists(credentials_file):
        logger.warning(f"Google credentials file '{credentials_file}' not found. Cannot sync to Google Sheets.")
        logger.warning("To enable Google Sheets sync, place a service account JSON file in the project root named 'credentials.json'.")
        return False
        
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        
        logger.info(f"Syncing {len(leads)} leads to Google Sheet ID: {sheet_id}")
        
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = service_account.Credentials.from_service_account_file(credentials_file, scopes=scopes)
        service = build('sheets', 'v4', credentials=creds)
        
        # Format values
        values = []
        # Header first
        values.append(HEADERS)
        
        for idx, lead in enumerate(leads, 1):
            values.append([
                idx,
                lead.get("school_name", ""),
                "",
                lead.get("institution_type", "Matriculation"),
                lead.get("social_media", "Inactive"),
                lead.get("area_name", ""),
                lead.get("atmosphere", "Good"),
                lead.get("website_url", ""),
                lead.get("contact_number", ""),
                lead.get("address", ""),
                lead.get("pincode", ""),
                lead.get("appearance", "Redesign"),
                lead.get("remarks", "")
            ])
            
        body = {'values': values}
        
        # Clear existing sheet first
        service.spreadsheets().values().clear(
            spreadsheetId=sheet_id,
            range="Sheet1!A1:M1000"
        ).execute()
        
        # Write new values
        result = service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range="Sheet1!A1",
            valueInputOption="RAW",
            body=body
        ).execute()
        
        logger.info(f"Google Sheet updated successfully: {result.get('updatedCells')} cells updated.")
        return True
        
    except Exception as e:
        logger.error(f"Google Sheets API Sync failed: {e}")
        return False
